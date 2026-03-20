import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os


# ── Color Palette
C_HEADER_BG  = "1A1A2E"
C_HEADER_FG  = "E8F4F0"
C_ROW_ALT    = "EAF7F4"
C_ROW_NORMAL = "FFFFFF"
C_SUCCESS    = "D5F5E3"
C_PARTIAL    = "FEF9E7"
C_FAILED     = "FADBD8"
C_BORDER     = "CACFD2"
C_TITLE      = "0E6655"


def _thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _conf_fill(confidence):
    if confidence == 1.0:
        return PatternFill("solid", fgColor=C_SUCCESS)
    elif confidence > 0:
        return PatternFill("solid", fgColor=C_PARTIAL)
    else:
        return PatternFill("solid", fgColor=C_FAILED)


def export_to_excel(data, fields, output_path="output.xlsx"):
    """
    Writes a professionally styled 3-sheet Excel file.
    Sheet 1  Extracted Data  full table with confidence colour coding
    Sheet 2  Summary         statistics overview
    Sheet 3  Chart           bar chart of extraction success
    Returns the output path.
    """

    if not data:
        data = []
    if not fields:
        fields = []

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — Extracted Data
    # ══════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Extracted Data"
    ws.sheet_view.showGridLines = False

    headers = ["#", "File"] + fields + ["Confidence", "Status"]
    hdr_fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    hdr_font  = Font(bold=True, color=C_HEADER_FG, size=11, name="Calibri")
    alt_fill  = PatternFill("solid", fgColor=C_ROW_ALT)
    norm_fill = PatternFill("solid", fgColor=C_ROW_NORMAL)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()
    ws.row_dimensions[1].height = 26

    total_full = total_partial = total_failed = 0

    for i, row in enumerate(data, 1):
        row_num    = i + 1
        confidence = float(row.get("confidence", 0))
        base_fill  = alt_fill if i % 2 == 0 else norm_fill

        if confidence == 1.0:
            total_full += 1;    status_text = "Complete"
        elif confidence > 0:
            total_partial += 1; status_text = "Partial"
        else:
            total_failed += 1;  status_text = "Failed"

        values = [i, row.get("file", f"file_{i}")]
        for f in fields:
            values.append(row.get(f, "—"))
        values.append(f"{int(confidence * 100)}%")
        values.append(status_text)

        last = len(values)
        for col_idx, val in enumerate(values, 1):
            cell           = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col_idx in (1, last, last-1) else "left")
            cell.border    = _thin_border()
            cell.font      = Font(size=10, name="Calibri")

            if col_idx == last - 1:          # Confidence
                cell.fill = _conf_fill(confidence)
                cell.font = Font(bold=True, size=10, name="Calibri")
            elif col_idx == last:            # Status
                cell.fill = _conf_fill(confidence)
                status_color = "155724" if confidence == 1.0 else "856404" if confidence > 0 else "721C24"
                cell.font = Font(bold=True, size=10, name="Calibri", color=status_color)
            else:
                cell.fill = base_fill

        ws.row_dimensions[row_num].height = 20

    col_widths = [6, 28] + [max(len(f) + 4, 16) for f in fields] + [14, 14]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 42)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — Summary
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 28

    title_font = Font(bold=True, size=15, color=C_TITLE,  name="Calibri")
    label_font = Font(bold=True, size=11, color="2C3E50", name="Calibri")
    value_font = Font(size=11,            color="1A1A2E", name="Calibri")
    meta_font  = Font(size=9, italic=True, color="7F8C8D", name="Calibri")

    total = len(data)
    pct = lambda n: f"{round(n/total*100)}%" if total else "0%"

    summary_rows = [
        ("ID-RSS — Extraction Summary",       "",                                        title_font, None),
        ("",                                  "",                                        None,       None),
        ("Generated on",    datetime.now().strftime("%d %b %Y, %I:%M %p"),              label_font, value_font),
        ("Generated by",    "ID-RSS  HackIndia 2026, KCC Institute",                    label_font, value_font),
        ("",                                  "",                                        None,       None),
        ("STATISTICS",                        "",                                        label_font, None),
        ("Total files processed",             total,                                     label_font, value_font),
        ("Fields extracted",                  ", ".join(fields) if fields else "-",      label_font, value_font),
        ("",                                  "",                                        None,       None),
        ("Fully extracted (100%)",            f"{total_full}  ({pct(total_full)})",      label_font, Font(bold=True, size=11, name="Calibri", color="1E8449")),
        ("Partial extractions",               f"{total_partial}  ({pct(total_partial)})",label_font, Font(bold=True, size=11, name="Calibri", color="B7770D")),
        ("Failed extractions",                f"{total_failed}  ({pct(total_failed)})",  label_font, Font(bold=True, size=11, name="Calibri", color="922B21")),
        ("",                                  "",                                        None,       None),
        ("NOTES",                             "",                                        label_font, None),
        ("Confidence = fields found / total", "",                                        meta_font,  None),
        ("A dash means field not in file",    "",                                        meta_font,  None),
    ]

    for r_idx, (label, value, l_font, v_font) in enumerate(summary_rows, 1):
        ca = ws2.cell(row=r_idx, column=1, value=label)
        cb = ws2.cell(row=r_idx, column=2, value=value)
        if l_font: ca.font = l_font
        if v_font and value != "": cb.font = v_font
        ws2.row_dimensions[r_idx].height = 20 if label else 8

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — Chart
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Chart")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 16

    for cell, val in [("A1","Status"), ("B1","Count")]:
        c = ws3[cell]
        c.value = val
        c.font  = Font(bold=True, name="Calibri", color=C_HEADER_FG)
        c.fill  = PatternFill("solid", fgColor=C_HEADER_BG)

    chart_data = [("Fully Extracted", total_full), ("Partial", total_partial), ("Failed", total_failed)]
    for row_i, (label, count) in enumerate(chart_data, 2):
        ws3.cell(row=row_i, column=1, value=label).font = Font(name="Calibri")
        ws3.cell(row=row_i, column=2, value=count).font = Font(name="Calibri")

    chart = BarChart()
    chart.type          = "col"
    chart.title         = "Extraction Results"
    chart.y_axis.title  = "Number of Files"
    chart.x_axis.title  = "Status"
    chart.style         = 10
    chart.width         = 16
    chart.height        = 10

    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=4)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws3.add_chart(chart, "D2")

    # ── Save
    out_dir = os.path.dirname(output_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_path)
    return output_path


# ── Quick test
if __name__ == "__main__":
    test_data = [
        {"file": "student_0001.docx", "Name": "Tanishk Bansal",  "Serial No.": "STU-2024-0001", "confidence": 1.0},
        {"file": "student_0002.docx", "Name": "Priya Sharma",    "Serial No.": "STU-2024-0002", "confidence": 1.0},
        {"file": "student_0003.docx", "Name": "Rahul Verma",     "Serial No.": "—",             "confidence": 0.5},
        {"file": "student_0004.docx", "Name": "—",               "Serial No.": "—",             "confidence": 0.0},
    ]
    path = export_to_excel(test_data, ["Name", "Serial No."], "test_output.xlsx")
    print(f"Excel created: {path}")
